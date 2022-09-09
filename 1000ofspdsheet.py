#!/usr/bin/env python
# coding: utf-8

# In[3]:


import openpyxl as xl
from openpyxl.chart import BarChart, Reference 
def process_workbook(filename):
    wb = xl.load_workbook('filename')
    sheet = wb['Sheet1']
    
    #now lets see how to make chart
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 8  #how we use float here 
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price
    
#reference class vlues object    
    values = Reference(sheet, 
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')  #lets see where we add our chart we add on e2 col,row


    wb.save('filename.xlsx')


# In[ ]:


#if u need multiple sheet then follow this 1000of spread sheet in seconds

