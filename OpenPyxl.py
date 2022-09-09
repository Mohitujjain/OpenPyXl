#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl as xl


# In[11]:


#when u need chart then u add this is 2nd way how we make chart 
from openpyxl.chart import BarChart, Reference   #module- chart,twolasses-BarChart,Reference


# In[2]:


wb = xl.load_workbook('transaction_id.xlsx')


# In[3]:


sheet = wb['Sheet1']


# In[4]:


cell = sheet['a1']


# In[5]:


cell = sheet.cell(1,1)


# In[6]:


print(cell.value)


# In[7]:


print(sheet.max_row)  #how many rows in sheet


# In[8]:


for row in range(1, sheet.max_row +1):   #if u (1,4)u get one to 3 means 123 if u use this u get exact
    print(row)


# In[9]:


#if u need particular value then use this:
for row in range(2, sheet.max_row +1):  #start with 2row
    cell = sheet.cell(row, 3)           #start with col3 and all value of that col3
    print(cell.value)


# In[10]:


#now lets start how we add new column with using formula of other column
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 8  #how we use float here 
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price
    
wb.save('transaction3.xlsx')


# In[12]:


#now lets see how to make chart
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 8  #how we use float here 
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price
    
#reference class vlues object    
values = Reference(sheet, 
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')  #lets see where we add our chart we add on e2 col,row


wb.save('transaction3.xlsx')


# In[ ]:




